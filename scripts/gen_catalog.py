# -*- coding: utf-8 -*-
"""Generate lib/catalog.data.ts from the client's Excel inventory."""
import openpyxl, json, re, hashlib
from collections import defaultdict, OrderedDict

SRC = r"C:/Users/welcome/Desktop/DelightStudio/items_Delight_Studio .xlsx"
OUT = r"C:/Users/welcome/Desktop/DelightStudio/lib/catalog.data.ts"

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb.active
H = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
def cell(r, n): return ws.cell(r, H[n]).value

rows = []
for r in range(2, ws.max_row + 1):
    nm = cell(r, "Item Name")
    if nm in (None, ""):
        continue
    rows.append({
        "name": str(nm).strip(),
        "cat": str(cell(r, "Category Name")),
        "price": cell(r, "Sales Rate") or 0,
        "stock": cell(r, "Available Stock") or 0,
    })

def sid(s):  # stable int hash
    return int(hashlib.md5(s.encode()).hexdigest(), 16)

# ---------- helpers ----------
def title(s):
    small = {"and": "&"}
    out = []
    for w in s.split():
        lw = w.lower()
        if lw in small:
            out.append(small[lw]); continue
        out.append(w.capitalize() if not w.isupper() or len(w) > 3 else w.capitalize())
    return " ".join(out)

SIZES = ["S", "M", "L", "XL", "XXL"]
SIZE_SUFFIX = re.compile(r"[-\s]+(XXXL|XXL|XL|L|M|S|\d{2,3})$", re.I)

def base_and_size(name):
    m = SIZE_SUFFIX.search(name)
    if m:
        return name[:m.start()].strip(" -"), m.group(1).upper()
    return name, None

# ---------- discount / mrp ----------
def price_mrp_off(price, seed):
    disc = [0.40, 0.45, 0.50, 0.55, 0.60][sid(seed) % 5]
    raw = price / (1 - disc)
    mrp = int(round(raw / 100.0)) * 100 - 1
    if mrp <= price:
        mrp = int(price * 1.6)
    off = round((mrp - price) / mrp * 100)
    return mrp, off

# ---------- images ----------
PIMG = ["p1", "p2", "p3", "p4", "p5", "p6"]
CATIMG = ["cat-kurta", "cat-anarkali", "cat-kurta-sets", "cat-skirt-sets",
          "cat-ethnic-dress", "cat-bottomwear", "cat-shop-all", "cat-nighty"]
def imgset(i):
    a = PIMG[i % 6]
    b = PIMG[(i + 2) % 6]
    c = CATIMG[i % len(CATIMG)]
    return [a, b, c]

COLORS = ["Ivory", "Burgundy", "Emerald", "Blush", "Indigo", "Rosewood", "Mustard",
          "Onyx", "Teal", "Coral", "Wine", "Sand", "Olive", "Saffron", "Mint"]
def pick_color(name, seed):
    for c in COLORS + ["White", "Black", "Blue", "Green", "Pink", "Grey", "Gray",
                       "Brown", "Skin", "Violet", "Rust", "Mustard", "Orange"]:
        if c.lower() in name.lower():
            return "Grey" if c in ("Gray",) else c
    return COLORS[sid(seed) % len(COLORS)]

def size_subset(seed, stock):
    patterns = [SIZES, ["S", "M", "L", "XL"], ["M", "L", "XL", "XXL"],
                ["S", "M", "L"], ["S", "M", "L", "XL", "XXL"], ["M", "L", "XL"]]
    return patterns[sid(seed) % len(patterns)]

def rating(seed):
    r = 4.2 + (sid(seed) % 8) / 10.0
    return round(min(r, 4.9), 1)
def reviews(seed):
    return 18 + sid(seed + "rev") % 300

# ---------- saree fabric detection ----------
SAREE_TYPES = [
    ("Kalamkari", ["kalamkari"]), ("Tissue", ["tissue"]), ("Kanchi Silk", ["pattu", "kubara"]),
    ("Silk", ["silk", "soft silk", "poly silk", "tussar", "tasar", "pochampally"]),
    ("Cotton", ["cotton", "malmal", "mul mul", "mul", "khadi", "kadhi", "handloom", "linen cotton"]),
    ("Linen", ["linen", "grosia"]), ("Organza", ["organza"]),
    ("Georgette", ["georgette"]), ("Chiffon", ["chiffon"]),
    ("Embroidery", ["embroidery", "cutwork", "aari", "handwork", "kathawork", "moti", "stone"]),
    ("Sequin", ["sequin", "jarkan", "sequin full"]),
    ("Digital Print", ["digital", "print", "ajrakh"]),
    ("Kerala", ["kerala"]), ("Satin", ["satin"]), ("Mirror", ["mirror"]),
]
def saree_type(name):
    n = name.lower()
    for label, keys in SAREE_TYPES:
        if any(k in n for k in keys):
            return label
    return "Fancy"

SAREE_PREFIX = re.compile(r"^(SU|AS|AM|TH|AY|PC|PDC|DF3|DF|KYI|SV|SC|TD|NE|PT|RRR|MA)\b\.?\s*", re.I)
SAREE_EXPAND = {"bdr": "Border", "selp": "Self", "rp": "", "putta": "Putta", "kadhi": "Khadi",
                "grosia": "Grosia", "sc": "", "pt": "", "aari": "Aari", "malmal": "Malmal"}
def saree_name(raw, stype):
    s = SAREE_PREFIX.sub("", raw)
    s = re.sub(r"\brp\b", "", s, flags=re.I)
    s = re.sub(r"[-\s]+\d+$", "", s).strip()          # trailing " - 2"
    s = re.sub(r"^\d+$", "", s).strip()               # names that are just "8"
    words = []
    for w in s.split():
        lw = w.lower()
        if lw in SAREE_EXPAND:
            e = SAREE_EXPAND[lw]
            if e: words.append(e)
        else:
            words.append(w.capitalize() if not w.isupper() else w.title())
    core = " ".join(words).strip()
    if not core:
        core = stype
    low = core.lower()
    if "saree" not in low and "pattu" not in low:
        core = f"{core} Saree"
    return re.sub(r"\s+", " ", core).strip()

# ---------- western ----------
def western_type(name):
    n = name.lower()
    if "crop top" in n: return "Crop Tops"
    if "gown" in n: return "Gowns"
    if "shirt" in n: return "Shirts"
    if any(k in n for k in ["dress", "midi", "maxi"]): return "Dresses"
    if any(k in n for k in ["pant", "trouser"]): return "Trousers"
    if any(k in n for k in ["2 piece", "co-ord", "coord"]): return "Co-ords"
    if "kurti" in n: return "Short Kurtis"
    return "Tops"
WEST_MAP = {
    "inner gray": "Grey Inner Top", "inner grey": "Grey Inner Top",
    "white short": "White Shorts", "top": "Essential Top", "top 1": "Essential Top I",
    "top 2": "Essential Top II", "top 3": "Essential Top III", "2 piece top": "Two-Piece Top Set",
    "peblum top": "Peplum Top", "womens pants": "Women's Trousers", "white top 1": "Classic White Top",
    "white top": "Classic White Top", "printed white": "Printed White Dress",
    "white and green": "White & Green Dress", "violet embroidery": "Violet Embroidery Dress",
    "shirt": "Classic Shirt", "design": "Designer Top",
}
def western_name(raw, wtype, price):
    key = raw.lower().strip()
    if key in WEST_MAP:
        return WEST_MAP[key]
    s = title(raw)
    s = re.sub(r"\s+", " ", s).strip()
    has_noun = any(k in s.lower() for k in
                   ["top", "shirt", "dress", "gown", "pant", "trouser", "short", "kurti", "tunic", "set"])
    if not has_noun:
        s = f"{s} Dress" if price and price > 1200 else f"{s} Top"
    return s

# ---------- pants ----------
def pant_type(name):
    n = name.lower()
    if "kurta pant" in n: return "Kurta Pants"
    if "palazzo" in n or n.startswith("p.pants"): return "Palazzo"
    if "baggy" in n: return "Baggy Jeans"
    if "straight" in n: return "Straight Jeans"
    if "jean" in n: return "Jeans"
    return "Casual Pants"
def pant_name(raw):
    n = raw
    n = re.sub(r"^MA\s+", "", n)
    n = re.sub(r"\b\d?P\b", "", n)          # 4P / 6P / 2P pocket codes
    n = re.sub(r"\bD indigo\b", "Dark Indigo", n, flags=re.I)
    n = re.sub(r"\bjeans?\b", "Jeans", n, flags=re.I)
    n = re.sub(r"\s+", " ", n).strip()
    two = bool(re.search(r"\b2$", n))
    n = re.sub(r"\s*2$", "", n).strip()
    if raw.strip() == "K.Pants": n = "Cotton Casual Pants"
    elif raw.strip() == "P.Pants": n = "Palazzo Pants"
    elif "jean" not in n.lower() and "pant" not in n.lower():
        n = n + " Jeans"
    n = title(n)
    if two:
        n = n + " II"
    return n

# ---------- hand maps for coded categories ----------
KURTI_MAP = {
    "FNAY": ("Fancy Embroidered Kurti", ["Embroidered"]),
    "SIY": ("Signature Straight Kurti", ["Straight"]),
    "FNCY": ("Fancy Georgette Kurti", ["Embroidered"]),
    "AVELEN": ("Avelen Cotton Kurti", ["Cotton"]),
    "RFT": ("Ruffle A-Line Kurti", ["A-Line"]),
    "AVAASA VATICAN": ("Avaasa Vatican Printed Kurti", ["Printed"]),
    "SATTIN": ("Satin Straight Kurti", ["Straight"]),
    "KURTI AV": ("Avaasa Everyday Kurti", ["Cotton"]),
    "AVAASHA KURTI": ("Avaasha Straight Kurti", ["Straight"]),
}
INNER_MAP = {
    "GLORIA PLAIN": ("Gloria Plain Camisole", ["Camisole"]),
    "LOVE ONE PRINT(3PC)": ("Love One Printed Camisole (3 Pc)", ["Camisole"]),
    "SPORTS BRASSIER": ("Sports Bra", ["Bra"]),
    "LIYA BRASSIER": ("Liya Everyday Bra", ["Bra"]),
    "REKHASLIPS WHITE": ("Rekha Slip \u2014 White", ["Slip"]),
    "REKHASLIPS SKIN": ("Rekha Slip \u2014 Skin", ["Slip"]),
    "REKHASLIPS BLACK": ("Rekha Slip \u2014 Black", ["Slip"]),
}
ANARKALI_MAP = {
    "FNCY": ("Fancy Flared Anarkali", ["Flared"]),
    "IRY 2": ("Ivory Georgette Anarkali", ["Georgette"]),
    "RAY": ("Regal Embroidered Anarkali", ["Embroidered"]),
    "AVELEN SET": ("Avelen Anarkali Set", ["Anarkali Set"]),
    "IYY": ("Ivory Embroidered Anarkali", ["Embroidered"]),
    "FCCY 2": ("Fancy Chiffon Anarkali II", ["Chiffon"]),
    "FCTY 2": ("Festive Flared Anarkali II", ["Flared"]),
    "FCTY": ("Festive Flared Anarkali", ["Flared"]),
    "FCCY": ("Fancy Chiffon Anarkali", ["Chiffon"]),
    "ICT": ("Ink Cotton Anarkali", ["Cotton"]),
}
ETHNIC_MAP = {
    "YHY": ("Handloom Ethnic Suit", ["Suit"]),
    "SRY": ("Serene Ethnic Kurta Set", ["Kurta Set"]),
    "ICY": ("Ivory Ethnic Suit", ["Suit"]),
    "HIY": ("Heritage Ethnic Suit", ["Suit"]),
}
SET3_MAP = {
    "IHY": ("Ivory 3-Piece Suit Set", ["3-Piece"]),
    "RAT": ("Radiant 3-Piece Suit Set", ["3-Piece"]),
    "FCTT": ("Festive 3-Piece Set", ["3-Piece"]),
    "FTHT": ("Floral 3-Piece Set", ["3-Piece"]),
}
SET2_MAP = {
    "AVAASA SET 2": ("Avaasa 2-Piece Co-ord II", ["2-Piece"]),
    "IRY 3": ("Ivory 2-Piece Set", ["2-Piece"]),
    "AVAASA SET": ("Avaasa 2-Piece Co-ord", ["2-Piece"]),
}
SINGLE_MAP = {
    "IRY": ("Ivory Single-Piece Kurta", ["Single"]),
    "SNY": ("Sunny Single-Piece Kurta", ["Single"]),
    "HHY": ("Heritage Single-Piece Kurta", ["Single"]),
}
FABRIC_BY_TYPE = {
    "Cotton": "Pure Cotton", "Silk": "Art Silk", "Kanchi Silk": "Kanchipuram Silk",
    "Tissue": "Tissue", "Linen": "Linen", "Organza": "Organza", "Georgette": "Georgette",
    "Chiffon": "Chiffon", "Kalamkari": "Cotton", "Embroidery": "Georgette", "Sequin": "Net",
    "Digital Print": "Poly Silk", "Kerala": "Cotton", "Satin": "Satin", "Mirror": "Cotton",
    "Fancy": "Art Silk", "Straight": "Cotton Blend", "A-Line": "Rayon", "Printed": "Rayon",
    "Embroidered": "Georgette", "Flared": "Georgette", "Chiffon ": "Chiffon",
    "Camisole": "Cotton", "Bra": "Cotton Blend", "Slip": "Satin", "Suit": "Cotton Blend",
    "Kurta Set": "Cotton", "3-Piece": "Georgette", "2-Piece": "Rayon", "Single": "Cotton Blend",
    "Anarkali Set": "Georgette", "Tops": "Cotton", "Crop Tops": "Cotton", "Shirts": "Cotton",
    "Dresses": "Crepe", "Gowns": "Georgette", "Trousers": "Cotton", "Co-ords": "Rayon",
    "Short Kurtis": "Rayon", "Jeans": "Denim", "Baggy Jeans": "Denim", "Straight Jeans": "Denim",
    "Palazzo": "Rayon", "Casual Pants": "Cotton",
}

# ---------- build products ----------
products = []
used_slugs = {}

def slugify(name):
    s = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    if s in used_slugs:
        used_slugs[s] += 1
        s = f"{s}-{used_slugs[s]}"
    else:
        used_slugs[s] = 1
    return s

def add(name, code, price, stock, cat_display, cat_slug, types, sizes, free, i):
    if not price or price <= 0:
        price = 499
    mrp, off = price_mrp_off(int(price), code + name)
    ptype = types[0] if types else "Fancy"
    fabric = FABRIC_BY_TYPE.get(ptype, "Art Silk")
    color = pick_color(name, code + name)
    slug = slugify(name)
    if free:
        desc = (f"A handpicked {ptype.lower()} saree in a {color.lower()} palette, woven for "
                f"effortless drape and all-day comfort. Finished with Delight Studio's signature "
                f"attention to detail \u2014 a graceful pick for festive and everyday wear.")
    else:
        desc = (f"A {ptype.lower()} {cat_display.lower()} in {color.lower()}, tailored from "
                f"{fabric.lower()} for an easy, flattering fit. Finished with Delight Studio's "
                f"signature attention to detail \u2014 effortless from day to evening.")
    tag = None
    if stock and stock >= 12: tag = "Bestseller"
    elif sid(code) % 7 == 0: tag = "New"
    elif off >= 58: tag = "Editor's Pick"
    products.append(OrderedDict([
        ("slug", slug), ("code", code), ("name", name),
        ("price", int(price)), ("mrp", mrp),
        ("image", f"/images/{imgset(i)[0]}.jpg"), ("hoverImage", f"/images/{imgset(i)[1]}.jpg"),
        ("images", [f"/images/{x}.jpg" for x in imgset(i)]),
        ("href", f"/product/{slug}"),
        ("category", cat_display), ("categorySlug", cat_slug),
        ("types", types), ("sizes", sizes), ("freeSize", free),
        ("stock", int(stock)), ("fabric", fabric), ("color", color),
        ("description", desc), ("rating", rating(code)), ("reviews", reviews(code)),
        ("tag", tag),
    ]))

idx = 0
# group rows by category
bycat = defaultdict(list)
for row in rows:
    bycat[row["cat"]].append(row)

# ----- SAREES (free size, one per row) -----
for row in bycat["sarees"]:
    st = saree_type(row["name"])
    nm = saree_name(row["name"], st)
    add(nm, row["name"], row["price"], row["stock"], "Sarees", "sarees",
        [st], [], True, idx); idx += 1

# ----- WESTERN (+ None items) -----
west_rows = list(bycat["Western"])
# fold None: White Top, design-s/m
none_rows = list(bycat["None"])
# collapse design-s / design-m -> Designer Top
design = [r for r in none_rows if r["name"].lower().startswith("design")]
if design:
    stk = sum(r["stock"] for r in design)
    pr = max((r["price"] for r in design), default=799)
    add("Designer Top", "design", pr, stk, "Western", "tops", ["Tops"],
        ["S", "M"], False, idx); idx += 1
for r in none_rows:
    if r["name"].lower().startswith("design"): continue
    wt = western_type(r["name"]); nm = western_name(r["name"], wt, r["price"])
    add(nm, r["name"], r["price"], r["stock"], "Western", "tops", [wt],
        size_subset(r["name"], r["stock"]), False, idx); idx += 1
for r in west_rows:
    wt = western_type(r["name"]); nm = western_name(r["name"], wt, r["price"])
    add(nm, r["name"], r["price"], r["stock"], "Western", "tops", [wt],
        size_subset(r["name"], r["stock"]), False, idx); idx += 1

# ----- KURTI (collapse size variants) -----
kgroups = OrderedDict()
for r in bycat["KURTI"]:
    base, sz = base_and_size(r["name"])
    kgroups.setdefault(base.upper(), {"base": base, "rows": []})["rows"].append((r, sz))
for baseU, g in kgroups.items():
    name, types = KURTI_MAP.get(baseU, (title(g["base"]) + " Kurti", ["Straight"]))
    stock = sum(rr["stock"] for rr, _ in g["rows"])
    price = max(rr["price"] for rr, _ in g["rows"])
    add(name, g["base"], price, stock, "Kurti", "kurtis", types, list(SIZES), False, idx); idx += 1

# ----- INNERWEAR (collapse) -----
igroups = OrderedDict()
for r in bycat["Innerwears"]:
    base, sz = base_and_size(r["name"])
    igroups.setdefault(base.upper(), {"base": base, "rows": []})["rows"].append(r)
for baseU, g in igroups.items():
    name, types = INNER_MAP.get(baseU, (title(g["base"]), ["Camisole"]))
    stock = sum(rr["stock"] for rr in g["rows"])
    prices = [rr["price"] for rr in g["rows"] if rr["price"]]
    price = min(prices) if prices else 199
    add(name, g["base"], price, stock, "Innerwear", "innerwear", types,
        list(SIZES), False, idx); idx += 1

# ----- PANTS -----
kurta_pants = [r for r in bycat["Pants"] if r["name"].lower().startswith("kurta pant")]
if kurta_pants:
    stk = sum(r["stock"] for r in kurta_pants)
    add("Kurta Pants", "Kurta Pants", 399, stk, "Pants", "pants", ["Kurta Pants"],
        ["L", "XL", "XXL"], False, idx); idx += 1
for r in bycat["Pants"]:
    if r["name"].lower().startswith("kurta pant"): continue
    pt = pant_type(r["name"]); nm = pant_name(r["name"])
    add(nm, r["name"], r["price"], r["stock"], "Pants", "pants", [pt],
        size_subset(r["name"], r["stock"]), False, idx); idx += 1

# ----- ANARKALI -----
for r in bycat["ANARKALI"]:
    name, types = ANARKALI_MAP.get(r["name"].upper(), (title(r["name"]) + " Anarkali", ["Flared"]))
    add(name, r["name"], r["price"], r["stock"], "Anarkali", "anarkalis", types,
        size_subset(r["name"], r["stock"]), False, idx); idx += 1

# ----- ETHNIC SETS -----
for r in bycat["ETHNIC"]:
    name, types = ETHNIC_MAP.get(r["name"].upper(), (title(r["name"]) + " Ethnic Suit", ["Suit"]))
    add(name, r["name"], r["price"], r["stock"], "Ethnic Set", "ethnic-sets", types,
        size_subset(r["name"], r["stock"]), False, idx); idx += 1

# ----- PIECE SETS (2 + 3 + single) -----
for r in bycat["3 PIECE SET"]:
    name, types = SET3_MAP.get(r["name"].upper(), (title(r["name"]) + " 3-Piece Set", ["3-Piece"]))
    add(name, r["name"], r["price"], r["stock"], "Piece Set", "piece-sets", types,
        size_subset(r["name"], r["stock"]), False, idx); idx += 1
for r in bycat["2 PIECE SET"]:
    name, types = SET2_MAP.get(r["name"].upper(), (title(r["name"]) + " 2-Piece Set", ["2-Piece"]))
    add(name, r["name"], r["price"], r["stock"], "Piece Set", "piece-sets", types,
        size_subset(r["name"], r["stock"]), False, idx); idx += 1
for r in bycat["SINGLE PC"]:
    name, types = SINGLE_MAP.get(r["name"].upper(), (title(r["name"]) + " Kurta", ["Single"]))
    add(name, r["name"], r["price"], r["stock"], "Piece Set", "piece-sets", types,
        size_subset(r["name"], r["stock"]), False, idx); idx += 1

# ---------- facets per leaf category slug ----------
facets = defaultdict(list)
cat_display_by_slug = {}
for p in products:
    cs = p["categorySlug"]; cat_display_by_slug[cs] = p["category"]
    for t in p["types"]:
        if t not in facets[cs]:
            facets[cs].append(t)

# ---------- newArrivals: in-stock, varied across categories ----------
new_slugs = []
seen_cat = set()
for p in products:
    if p["stock"] and p["categorySlug"] not in seen_cat:
        new_slugs.append(p["slug"]); seen_cat.add(p["categorySlug"])
    if len(new_slugs) >= 6: break
for p in products:
    if len(new_slugs) >= 8: break
    if p["stock"] and p["slug"] not in new_slugs:
        new_slugs.append(p["slug"])

# ---------- emit TS ----------
def js(v):
    return json.dumps(v, ensure_ascii=False)

lines = []
lines.append("// AUTO-GENERATED from items_Delight_Studio.xlsx — do not edit by hand.")
lines.append("// Regenerate via scripts/gen_catalog.py. `code` = original Excel item name")
lines.append("// (kept so the backend can match rows). Names/MRP are presentation-only.")
lines.append('import type { CatalogProduct } from "./data";')
lines.append("")
lines.append("export const CATALOG: CatalogProduct[] = [")
for p in products:
    fields = []
    for k, v in p.items():
        if v is None:
            continue
        fields.append(f"{k}: {js(v)}")
    lines.append("  { " + ", ".join(fields) + " },")
lines.append("];")
lines.append("")
lines.append("export const CATEGORY_FACETS: Record<string, string[]> = " +
             json.dumps(dict(facets), ensure_ascii=False, indent=2) + ";")
lines.append("")
lines.append("export const NEW_ARRIVAL_SLUGS: string[] = " + js(new_slugs) + ";")
lines.append("")
lines.append("export const CATEGORY_DISPLAY: Record<string, string> = " +
             json.dumps(cat_display_by_slug, ensure_ascii=False, indent=2) + ";")
lines.append("")

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))

# ---------- report ----------
print("PRODUCTS:", len(products))
cc = defaultdict(int)
for p in products: cc[p["categorySlug"]] += 1
for k in sorted(cc): print(f"  {k:14} {cc[k]:3}  facets={facets[k]}")
print("newArrivals:", new_slugs)
print("\nSample names:")
for p in products[:3] + products[92:96] + products[-8:]:
    print(f"  [{p['categorySlug']:11}] {p['name'][:34]:34} Rs{p['price']}  MRP{p['mrp']} "
          f"({round((p['mrp']-p['price'])/p['mrp']*100)}% off) stock={p['stock']} free={p['freeSize']}")
